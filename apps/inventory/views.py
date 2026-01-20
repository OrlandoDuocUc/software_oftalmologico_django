from __future__ import annotations

import json
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt

from apps.sales.services import SaleService

from .models import Proveedor
from .services import ProductService, PurchaseService
from openpyxl import Workbook


product_service = ProductService()
sale_service = SaleService()
purchase_service = PurchaseService()


def _require_admin(request):
    if request.session.get("rol") != "Administrador":
        messages.error(request, "Acceso no autorizado.")
        return redirect("user_html:login")
    return None


def _parse_date(value: str | None):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def _clean_int(value: str | None) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _clean_decimal(value: str | None) -> Decimal:
    try:
        return Decimal(value)
    except (TypeError, ValueError, ArithmeticError):
        return Decimal(0)


def _clean_date(value: str | None):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _attach_precio_iva(productos):
    iva_factor = Decimal("1.15")
    for p in productos:
        base = p.costo_unitario or Decimal(0)
        p.precio_iva = (base * iva_factor).quantize(Decimal("0.01"))
    return productos


@login_required
def dashboard(request):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response

    productos = list(product_service.list_products())
    _attach_precio_iva(productos)
    stock_total = sum(p.cantidad or 0 for p in productos)
    productos_bajo_stock = [p for p in productos if (p.cantidad or 0) <= 10]
    ventas_recientes = list(sale_service.get_all_sales_with_details()[:5])

    return render(
        request,
        "dashboard.html",
        {
            "stock_total_productos": stock_total,
            "productos_bajo_stock": productos_bajo_stock,
            "total_ventas_hoy": 0,
            "ventas_recientes": ventas_recientes,
            "pagos_pendientes": 0,
        },
    )


@login_required
def productos(request):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response

    if request.method == "POST":
        try:
            data = {
                "fecha": _parse_date(request.POST.get("fecha")),
                "nombre": request.POST.get("nombre"),
                "distribuidor": request.POST.get("distribuidor"),
                "marca": request.POST.get("marca"),
                "material": request.POST.get("material"),
                "tipo_armazon": request.POST.get("tipo_armazon"),
                "codigo": request.POST.get("codigo"),
                "diametro_1": request.POST.get("diametro_1"),
                "diametro_2": request.POST.get("diametro_2"),
                "color": request.POST.get("color"),
                "cantidad": _clean_int(request.POST.get("cantidad")),
                "costo_unitario": _clean_decimal(request.POST.get("costo_unitario")),
                "descripcion": request.POST.get("descripcion"),
                "estado": True,
            }
            product_service.create_product(data)
            messages.success(request, "Producto creado exitosamente.")
            return redirect("product_html:productos")
        except Exception as exc:
            messages.error(request, f"Ocurrió un error al crear el producto: {exc}")

    products = list(product_service.list_products())
    _attach_precio_iva(products)
    return render(request, "productos.html", {"products": products})


@login_required
def productos_eliminados(request):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response
    deleted = list(product_service.list_products(include_deleted=True, only_deleted=True))
    _attach_precio_iva(deleted)
    return render(request, "productos.html", {"products": [], "deleted_products": deleted})


@login_required
def editar_producto(request, product_id: int):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response
    if request.method != "POST":
        return redirect("product_html:productos")
    data = {
        "fecha": _parse_date(request.POST.get("fecha")),
        "nombre": request.POST.get("nombre"),
        "distribuidor": request.POST.get("distribuidor"),
        "marca": request.POST.get("marca"),
        "material": request.POST.get("material"),
        "tipo_armazon": request.POST.get("tipo_armazon"),
        "codigo": request.POST.get("codigo"),
        "diametro_1": request.POST.get("diametro_1"),
        "diametro_2": request.POST.get("diametro_2"),
        "color": request.POST.get("color"),
        "cantidad": request.POST.get("cantidad"),
        "costo_unitario": request.POST.get("costo_unitario"),
        "descripcion": request.POST.get("descripcion"),
    }
    product_service.update_product(product_id, data)
    messages.success(request, "Producto actualizado correctamente.")
    return redirect("product_html:productos")


@login_required
def eliminar_producto(request, product_id: int):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response
    if request.method != "POST":
        return redirect("product_html:productos")
    success = product_service.delete_product(product_id)
    if success:
        messages.success(request, "Producto eliminado físicamente.")
    else:
        messages.warning(
            request,
            "El producto no se eliminó físicamente (tiene ventas asociadas) y fue desactivado.",
        )
    return redirect("product_html:productos")


@login_required
def restaurar_producto(request, product_id: int):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response
    if request.method != "POST":
        return redirect("product_html:productos_eliminados")
    product_service.restore_product(product_id)
    messages.success(request, "Producto restaurado exitosamente.")
    return redirect("product_html:productos_eliminados")


@login_required
@csrf_exempt
def exportar_inventario_excel(request):
    """
    Genera un XLSX con los productos.
    - Si el frontend envía un JSON con "productos" (filtrados), se usa tal cual.
    - Si no, se exporta todo el inventario desde la base de datos.
    """
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response

    productos_json = []
    if request.body:
        try:
            payload = json.loads(request.body.decode("utf-8"))
            productos_json = payload.get("productos") or []
        except json.JSONDecodeError:
            productos_json = []

    if not productos_json:
        productos = product_service.list_products()
        productos_json = [
            {
                "fecha": p.fecha.strftime("%d/%m/%Y") if p.fecha else "",
                "nombre": p.nombre,
                "distribuidor": p.distribuidor or "",
                "marca": p.marca or "",
                "material": p.material or "",
                "tipo_armazon": p.tipo_armazon or "",
                "codigo": p.codigo or "",
                "diametro_1": p.diametro_1 or "",
                "diametro_2": p.diametro_2 or "",
                "color": p.color or "",
                "cantidad": p.cantidad or 0,
                "costo_unitario": float(p.costo_unitario or 0),
                "costo_total": float(p.costo_total or 0),
                "costo_venta_1": float(p.costo_venta_1 or 0),
                "costo_venta_2": float(p.costo_venta_2 or 0),
                "estado": "Activo" if p.estado else "Inactivo",
            }
            for p in productos
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = [
        "Fecha",
        "Nombre",
        "Distribuidor",
        "Marca",
        "Material",
        "Tipo Armazón",
        "Código",
        "Diámetro 1",
        "Diámetro 2",
        "Color",
        "Cantidad",
        "Costo Unitario",
        "Costo Unitario + IVA",
        "Costo Total",
        "Venta 1",
        "Venta 2",
        "Estado",
    ]
    ws.append(headers)

    for item in productos_json:
        base_cost = Decimal(str(item.get("costo_unitario", 0) or 0))
        costo_iva = (base_cost * Decimal("1.15")).quantize(Decimal("0.01"))
        ws.append(
            [
                item.get("fecha", ""),
                item.get("nombre", ""),
                item.get("distribuidor", ""),
                item.get("marca", ""),
                item.get("material", ""),
                item.get("tipo_armazon", ""),
                item.get("codigo", ""),
                item.get("diametro_1", ""),
                item.get("diametro_2", ""),
                item.get("color", ""),
                item.get("cantidad", 0),
                item.get("costo_unitario", 0),
                float(costo_iva),
                item.get("costo_total", 0),
                item.get("costo_venta_1", 0),
                item.get("costo_venta_2", 0),
                item.get("estado", "Activo"),
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return response


@login_required
def compras(request):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response

    proveedores = Proveedor.objects.filter(estado=True).order_by("razon_social")
    productos = list(product_service.list_products())
    compras_list = list(purchase_service.list_purchases())
    for compra in compras_list:
        compra.iva_total_calculado = (compra.iva_15 or Decimal(0)) + (compra.iva_5 or Decimal(0))

    if request.method == "POST":
        try:
            header = {
                "proveedor_id": _clean_int(request.POST.get("proveedor_id")),
                "numero_factura": request.POST.get("numero_factura"),
                "ruc_ci": request.POST.get("ruc_ci"),
                "fecha_pedido": _clean_date(request.POST.get("fecha_pedido")),
                "fecha_pago": _clean_date(request.POST.get("fecha_pago")),
                "forma_pago": request.POST.get("forma_pago"),
                "plazo_pago": request.POST.get("plazo_pago"),
                "notas": request.POST.get("notas"),
                "abono": _clean_decimal(request.POST.get("abono")),
                "elaborado_codigo": request.POST.get("elaborado_codigo"),
                "elaborado_nombre": request.POST.get("elaborado_nombre"),
                "autorizado_codigo": request.POST.get("autorizado_codigo"),
                "autorizado_nombre": request.POST.get("autorizado_nombre"),
                "recibido_codigo": request.POST.get("recibido_codigo"),
                "recibido_nombre": request.POST.get("recibido_nombre"),
                "estado": request.POST.get("estado") or "borrador",
            }

            detalles = []
            productos_ids = request.POST.getlist("detalle_producto_id[]")
            marcas = request.POST.getlist("detalle_marca[]")
            codigos = request.POST.getlist("detalle_codigo[]")
            descripciones = request.POST.getlist("detalle_descripcion[]")
            cantidades = request.POST.getlist("detalle_cantidad[]")
            precios = request.POST.getlist("detalle_precio_unitario[]")
            tarifas = request.POST.getlist("detalle_tarifa_iva[]")
            descuentos = request.POST.getlist("detalle_descuento[]")
            tipos = request.POST.getlist("detalle_tipo[]")
            nuevos_nombres = request.POST.getlist("detalle_nuevo_nombre[]")
            nuevos_distribuidores = request.POST.getlist("detalle_nuevo_distribuidor[]")
            nuevos_rubros = request.POST.getlist("detalle_nuevo_rubro[]")
            nuevos_marcas = request.POST.getlist("detalle_nuevo_marca[]")
            nuevos_codigos = request.POST.getlist("detalle_nuevo_codigo[]")
            nuevos_materiales = request.POST.getlist("detalle_nuevo_material[]")
            nuevos_tipos_armazon = request.POST.getlist("detalle_nuevo_tipo_armazon[]")
            nuevos_diametro_1 = request.POST.getlist("detalle_nuevo_diametro_1[]")
            nuevos_diametro_2 = request.POST.getlist("detalle_nuevo_diametro_2[]")
            nuevos_colores = request.POST.getlist("detalle_nuevo_color[]")

            filas_detalle = len(productos_ids)
            for idx in range(filas_detalle):
                tipo = tipos[idx] if idx < len(tipos) else "existente"
                prod_id_raw = productos_ids[idx] if idx < len(productos_ids) else ""
                cantidad = _clean_decimal(cantidades[idx] if idx < len(cantidades) else None)
                if cantidad <= 0:
                    continue

                detalle = {
                    "marca": marcas[idx] if idx < len(marcas) else "",
                    "codigo": codigos[idx] if idx < len(codigos) else "",
                    "descripcion": descripciones[idx] if idx < len(descripciones) else "",
                    "cantidad": cantidad,
                    "precio_unitario": _clean_decimal(precios[idx] if idx < len(precios) else None),
                    "tarifa_iva": _clean_decimal(tarifas[idx] if idx < len(tarifas) else None),
                    "descuento": _clean_decimal(descuentos[idx] if idx < len(descuentos) else None),
                }

                if tipo == "nuevo":
                    nombre_nuevo = nuevos_nombres[idx].strip() if idx < len(nuevos_nombres) else ""
                    rubro_nuevo = nuevos_rubros[idx].strip() if idx < len(nuevos_rubros) else ""
                    if not nombre_nuevo or not rubro_nuevo:
                        raise ValueError("Completa los datos del producto nuevo.")

                    detalle["producto_id"] = None
                    detalle["nuevo_producto"] = {
                        "nombre": nombre_nuevo,
                        "rubro": rubro_nuevo,
                        "distribuidor": nuevos_distribuidores[idx].strip() if idx < len(nuevos_distribuidores) else "",
                        "marca": nuevos_marcas[idx].strip() if idx < len(nuevos_marcas) else "",
                        "codigo": nuevos_codigos[idx].strip() if idx < len(nuevos_codigos) else "",
                        "material": nuevos_materiales[idx].strip() if idx < len(nuevos_materiales) else "",
                        "tipo_armazon": nuevos_tipos_armazon[idx].strip() if idx < len(nuevos_tipos_armazon) else "",
                        "diametro_1": nuevos_diametro_1[idx].strip() if idx < len(nuevos_diametro_1) else "",
                        "diametro_2": nuevos_diametro_2[idx].strip() if idx < len(nuevos_diametro_2) else "",
                        "color": nuevos_colores[idx].strip() if idx < len(nuevos_colores) else "",
                    }
                else:
                    if not prod_id_raw:
                        continue
                    detalle["producto_id"] = _clean_int(prod_id_raw)

                detalles.append(detalle)

            if not detalles:
                messages.error(request, "Agrega al menos un detalle de producto.")
            else:
                proveedor_obj = None
                if header.get("proveedor_id"):
                    proveedor_obj = Proveedor.objects.filter(proveedor_id=header["proveedor_id"]).first()

                for det in detalles:
                    nuevo_payload = det.pop("nuevo_producto", None)
                    if nuevo_payload is None:
                        continue

                    product_data = {
                        "fecha": datetime.now(),
                        "nombre": nuevo_payload["nombre"],
                        "rubro": nuevo_payload["rubro"],
                        "distribuidor": nuevo_payload.get("distribuidor") or (proveedor_obj.razon_social if proveedor_obj else None),
                        "marca": nuevo_payload["marca"],
                        "material": nuevo_payload["material"],
                        "tipo_armazon": nuevo_payload["tipo_armazon"],
                        "codigo": nuevo_payload["codigo"],
                        "diametro_1": nuevo_payload.get("diametro_1"),
                        "diametro_2": nuevo_payload.get("diametro_2"),
                        "color": nuevo_payload["color"],
                        "cantidad": int(det["cantidad"]),
                        "costo_unitario": det["precio_unitario"],
                        "descripcion": det.get("descripcion"),
                        "estado": True,
                    }
                    nuevo_producto = product_service.create_product(product_data)
                    det["producto_id"] = nuevo_producto.producto_id

                    if not det.get("marca"):
                        det["marca"] = nuevo_producto.marca or ""
                    if not det.get("codigo"):
                        det["codigo"] = nuevo_producto.codigo or ""

                compra = purchase_service.create_purchase(header, detalles)
                messages.success(request, f"Compra #{compra.compra_id} registrada correctamente.")
                return redirect("product_html:compras")
        except Exception as exc:
            messages.error(request, f"Error al registrar la compra: {exc}")

    return render(
        request,
        "compras.html",
        {
            "proveedores": proveedores,
            "productos": productos,
            "compras": compras_list,
            "today": datetime.now().strftime("%Y-%m-%d"),
        },
    )


@login_required
def exportar_compras_excel(request):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response

    compras = purchase_service.list_purchases()

    wb = Workbook()
    ws_cab = wb.active
    ws_cab.title = "Compras"

    ws_cab.append(
        [
            "ID",
            "Proveedor",
            "Factura",
            "RUC/CI",
            "Fecha pedido",
            "Fecha pago",
            "Subtotal",
            "IVA",
            "Total pagar",
            "Abono",
            "Saldo",
            "Estado",
        ]
    )

    ws_det = wb.create_sheet(title="Detalles")
    ws_det.append(
        [
            "Compra ID",
            "Producto",
            "Marca",
            "Código",
            "Descripción",
            "Cantidad",
            "Precio unitario",
            "Tarifa IVA",
            "Descuento",
            "Valor total",
        ]
    )

    for c in compras:
        iva_total = (c.iva_15 or 0) + (c.iva_5 or 0)
        ws_cab.append(
            [
                c.compra_id,
                c.proveedor.razon_social if c.proveedor else "",
                c.numero_factura or "",
                c.ruc_ci or "",
                c.fecha_pedido.strftime("%d/%m/%Y") if c.fecha_pedido else "",
                c.fecha_pago.strftime("%d/%m/%Y") if c.fecha_pago else "",
                float(c.subtotal_general or 0),
                float(iva_total),
                float(c.total_pagar or 0),
                float(c.abono or 0),
                float(c.saldo or 0),
                c.estado or "",
            ]
        )
        for d in c.detalles.all():
            ws_det.append(
                [
                    c.compra_id,
                    d.producto.nombre if d.producto else "",
                    d.marca or "",
                    d.codigo or "",
                    d.descripcion or "",
                    d.cantidad or 0,
                    float(d.precio_unitario or 0),
                    float(d.tarifa_iva or 0),
                    float(d.descuento or 0),
                    float(d.valor_total or 0),
                ]
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"compras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def lista_proveedores(request):
    redirect_response = _require_admin(request)
    if redirect_response:
        return redirect_response
    proveedores = Proveedor.objects.all()
    return render(request, "proveedores.html", {"proveedores": proveedores})
